"""
Excel generator for creating formatted spreadsheets with multiple sheets and styling.
"""

import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional, Any, Union
from dataclasses import dataclass
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.utils.logger import logger
from src.utils.exceptions import FileProcessingError


@dataclass
class ExcelStyle:
    """Excel styling configuration."""
    
    header_font: Font = None
    header_fill: PatternFill = None
    data_font: Font = None
    border: Border = None
    alignment: Alignment = None
    
    def __post_init__(self):
        if self.header_font is None:
            self.header_font = Font(bold=True, color="FFFFFF")
        
        if self.header_fill is None:
            self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        if self.data_font is None:
            self.data_font = Font(color="000000")
        
        if self.border is None:
            thin_border = Side(border_style="thin", color="000000")
            self.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
        
        if self.alignment is None:
            self.alignment = Alignment(horizontal="left", vertical="center")


@dataclass
class ChartConfig:
    """Chart configuration."""
    
    chart_type: str  
    title: str
    data_range: str
    position: str = "E2"  
    width: int = 15
    height: int = 10
    
    def create_chart(self):
        """Create chart object based on type."""
        if self.chart_type.lower() == 'bar':
            chart = BarChart()
        elif self.chart_type.lower() == 'line':
            chart = LineChart()
        elif self.chart_type.lower() == 'pie':
            chart = PieChart()
        else:
            chart = BarChart()  
        
        chart.title = self.title
        chart.width = self.width
        chart.height = self.height
        
        return chart


@dataclass
class SheetConfig:
    """Configuration for a single Excel sheet."""
    
    name: str
    data: pd.DataFrame
    style: Optional[ExcelStyle] = None
    auto_filter: bool = True
    freeze_panes: Optional[str] = "A2"  
    column_widths: Optional[Dict[str, int]] = None
    charts: Optional[List[ChartConfig]] = None
    
    def __post_init__(self):
        if self.style is None:
            self.style = ExcelStyle()
        
        if self.charts is None:
            self.charts = []


class ExcelGenerator:
    """Generator for creating formatted Excel files."""
    
    def __init__(self):
        self.workbook = None
        self.default_style = ExcelStyle()
        logger.info("ExcelGenerator initialized")
    
    def create_workbook(self, 
                       sheets: List[SheetConfig],
                       output_path: Union[str, Path],
                       metadata: Optional[Dict[str, str]] = None) -> Path:
        """Create a complete Excel workbook with multiple sheets."""
        output_path = Path(output_path)
        
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            self.workbook = openpyxl.Workbook()
            
            default_sheet = self.workbook.active
            self.workbook.remove(default_sheet)
            
            for sheet_config in sheets:
                self._create_sheet(sheet_config)
            
            if metadata:
                self._set_workbook_metadata(metadata)
            
            self.workbook.save(output_path)
            
            logger.info(f"Excel workbook created: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to create Excel workbook: {e}")
            raise FileProcessingError(f"Excel generation failed: {e}")
    
    def _create_sheet(self, config: SheetConfig) -> None:
        """Create a single worksheet."""
        try:
            worksheet = self.workbook.create_sheet(title=config.name)
            
            if config.data.empty:
                logger.warning(f"Empty data for sheet: {config.name}")
                return
            
            self._populate_sheet_data(worksheet, config)
            
            self._apply_styling(worksheet, config)
            
            if config.auto_filter:
                self._add_auto_filter(worksheet, config.data)
            
            if config.freeze_panes:
                worksheet.freeze_panes = config.freeze_panes
            
            if config.column_widths:
                self._set_column_widths(worksheet, config.column_widths)
            else:
                self._auto_adjust_column_widths(worksheet, config.data)
            
            for chart_config in config.charts:
                self._add_chart(worksheet, chart_config, config.data)
            
            logger.debug(f"Sheet created: {config.name}")
            
        except Exception as e:
            logger.error(f"Failed to create sheet {config.name}: {e}")
            raise
    
    def _populate_sheet_data(self, worksheet, config: SheetConfig) -> None:
        """Populate worksheet with data."""
        for r in dataframe_to_rows(config.data, index=False, header=True):
            worksheet.append(r)
    
    def _apply_styling(self, worksheet, config: SheetConfig) -> None:
        """Apply styling to the worksheet."""
        try:
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = config.style.header_font
                cell.fill = config.style.header_fill
                cell.border = config.style.border
                cell.alignment = config.style.alignment
            
            for row in range(2, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = config.style.data_font
                    cell.border = config.style.border
                    cell.alignment = config.style.alignment
            
        except Exception as e:
            logger.warning(f"Failed to apply styling: {e}")
    
    def _add_auto_filter(self, worksheet, data: pd.DataFrame) -> None:
        """Add auto-filter to the data range."""
        if not data.empty:
            worksheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(data.columns))}{len(data) + 1}"
    
    def _set_column_widths(self, worksheet, column_widths: Dict[str, int]) -> None:
        """Set specific column widths."""
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width
    
    def _auto_adjust_column_widths(self, worksheet, data: pd.DataFrame) -> None:
        """Auto-adjust column widths based on content."""
        try:
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  
                worksheet.column_dimensions[column_letter].width = max(adjusted_width, 10)
                
        except Exception as e:
            logger.warning(f"Failed to auto-adjust column widths: {e}")
    
    def _add_chart(self, worksheet, chart_config: ChartConfig, data: pd.DataFrame) -> None:
        """Add chart to worksheet."""
        try:
            chart = chart_config.create_chart()
            
            if chart_config.data_range:
                data_ref = Reference(worksheet, range_string=chart_config.data_range)
                chart.add_data(data_ref, titles_from_data=True)
            else:
                min_col = 1
                min_row = 2  
                max_col = len(data.columns)
                max_row = len(data) + 1
                
                data_ref = Reference(worksheet, min_col=min_col, min_row=min_row, 
                                   max_col=max_col, max_row=max_row)
                chart.add_data(data_ref, titles_from_data=True)
            
            worksheet.add_chart(chart, chart_config.position)
            
        except Exception as e:
            logger.warning(f"Failed to add chart: {e}")
    
    def _set_workbook_metadata(self, metadata: Dict[str, str]) -> None:
        """Set workbook metadata properties."""
        props = self.workbook.properties
        
        props.title = metadata.get('title', 'Generated Report')
        props.subject = metadata.get('subject', 'Data Analysis Report')
        props.creator = metadata.get('creator', 'Web Scraper AI')
        props.description = metadata.get('description', 'Automatically generated Excel report')
        props.created = datetime.now()
    
    def create_summary_sheet(self, 
                           data_sheets: List[SheetConfig],
                           summary_stats: Optional[Dict[str, Any]] = None) -> SheetConfig:
        """Create a summary sheet with overview of all data."""
        try:
            summary_data = []
            
            for sheet in data_sheets:
                if not sheet.data.empty:
                    summary_data.append({
                        'Sheet Name': sheet.name,
                        'Total Rows': len(sheet.data),
                        'Total Columns': len(sheet.data.columns),
                        'Data Types': ', '.join(sheet.data.dtypes.astype(str).unique()),
                        'Memory Usage (KB)': round(sheet.data.memory_usage(deep=True).sum() / 1024, 2),
                        'Missing Values': sheet.data.isnull().sum().sum(),
                        'Last Updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    })
            
            if summary_stats:
                for key, value in summary_stats.items():
                    summary_data.append({
                        'Sheet Name': f"STAT: {key}",
                        'Total Rows': value if isinstance(value, (int, float)) else str(value),
                        'Total Columns': '',
                        'Data Types': '',
                        'Memory Usage (KB)': '',
                        'Missing Values': '',
                        'Last Updated': ''
                    })
            
            summary_df = pd.DataFrame(summary_data)
            
            summary_style = ExcelStyle()
            summary_style.header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            
            return SheetConfig(
                name="Summary",
                data=summary_df,
                style=summary_style,
                freeze_panes="A2"
            )
            
        except Exception as e:
            logger.error(f"Failed to create summary sheet: {e}")
            return SheetConfig(name="Summary", data=pd.DataFrame())
    
    def create_pivot_table_sheet(self, 
                                data: pd.DataFrame,
                                index_cols: List[str],
                                value_cols: List[str],
                                aggfunc: str = 'sum') -> SheetConfig:
        """Create a sheet with pivot table."""
        try:
            pivot_table = pd.pivot_table(
                data,
                index=index_cols,
                values=value_cols,
                aggfunc=aggfunc,
                fill_value=0
            )
            
            pivot_table = pivot_table.reset_index()
            
            pivot_style = ExcelStyle()
            pivot_style.header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
            
            return SheetConfig(
                name="Pivot Analysis",
                data=pivot_table,
                style=pivot_style
            )
            
        except Exception as e:
            logger.error(f"Failed to create pivot table: {e}")
            return SheetConfig(name="Pivot Analysis", data=pd.DataFrame())
    
    def add_data_validation(self, 
                          worksheet,
                          cell_range: str,
                          validation_type: str,
                          formula1: str,
                          formula2: Optional[str] = None) -> None:
        """Add data validation to a cell range."""
        try:
            from openpyxl.worksheet.datavalidation import DataValidation
            
            dv = DataValidation(
                type=validation_type,
                formula1=formula1,
                formula2=formula2,
                allow_blank=True
            )
            
            dv.add(cell_range)
            worksheet.add_data_validation(dv)
            
        except Exception as e:
            logger.warning(f"Failed to add data validation: {e}")
    
    def create_dashboard_sheet(self, 
                             charts: List[ChartConfig],
                             title: str = "Dashboard") -> SheetConfig:
        """Create a dashboard sheet with multiple charts."""
        dashboard_df = pd.DataFrame({
            'Dashboard': [title],
            'Created': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        })
        
        dashboard_style = ExcelStyle()
        dashboard_style.header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
        
        return SheetConfig(
            name="Dashboard",
            data=dashboard_df,
            style=dashboard_style,
            charts=charts,
            freeze_panes=None
        )
    
    def export_dataframes_to_excel(self, 
                                  dataframes: Dict[str, pd.DataFrame],
                                  output_path: Union[str, Path],
                                  include_summary: bool = True,
                                  include_charts: bool = False) -> Path:
        """Quick export multiple DataFrames to Excel."""
        try:
            sheets = []
            
            for name, df in dataframes.items():
                if not df.empty:
                    sheet_config = SheetConfig(name=name, data=df)
                    
                    if include_charts and len(df.select_dtypes(include=['number']).columns) > 0:
                        chart = ChartConfig(
                            chart_type='bar',
                            title=f'{name} Overview',
                            data_range='',
                            position='F2'
                        )
                        sheet_config.charts = [chart]
                    
                    sheets.append(sheet_config)
            
            if include_summary and sheets:
                summary_sheet = self.create_summary_sheet(sheets)
                sheets.insert(0, summary_sheet)
            
            return self.create_workbook(sheets, output_path)
            
        except Exception as e:
            logger.error(f"Failed to export DataFrames to Excel: {e}")
            raise FileProcessingError(f"Excel export failed: {e}")